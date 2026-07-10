"""Tests para worldclass_scraper.modules.exporters.

Cubre:
  - ExcelExporter: csv, xlsx, sanitize, truncate
  - ExportOrchestrator: partial (no-op), site_reports, combined
"""
import asyncio
import csv
import os
import pytest

from worldclass_scraper.modules.exporters import ExcelExporter, ExportOrchestrator


# ── fixtures ──────────────────────────────────────────────────────────────────

ROWS = [
    {'Sede': 'WCG - GUAYAQUIL', 'Estado_Contrato': 'CASH',  'Nombre_Titular': 'Ana'},
    {'Sede': 'WCG - GUAYAQUIL', 'Estado_Contrato': 'PROCE', 'Nombre_Titular': 'Luis'},
    {'Sede': 'WC - SANTO DOMINGO', 'Estado_Contrato': 'CASH', 'Nombre_Titular': 'María'},
]

SITIO = {
    'nombre': 'worldclass',
    'excel_template': 'contratos_{SEDE}_{ESTADO}.xlsx',
    'sheet_names': ['CASH', 'PROCE'],
    'filtros': {'sede': 'WCG - GUAYAQUIL'},
}


# ── ExcelExporter ─────────────────────────────────────────────────────────────

class TestExcelExporterCsv:
    def test_creates_file(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        path = exp.export_csv(ROWS, str(tmp_path / 'out.csv'))
        assert os.path.exists(path)

    def test_file_has_expected_rows(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        path = exp.export_csv(ROWS, str(tmp_path / 'out.csv'))
        with open(path, encoding='utf-8-sig') as fh:
            reader = list(csv.DictReader(fh))
        assert len(reader) == 3

    def test_relative_path_resolves_to_output_dir(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        path = exp.export_csv(ROWS, 'resultado.csv')
        assert path == str(tmp_path / 'resultado.csv')
        assert os.path.exists(path)

    def test_empty_rows_creates_file(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        path = exp.export_csv([], str(tmp_path / 'empty.csv'))
        assert os.path.exists(path)


class TestExcelExporterXlsx:
    def test_creates_xlsx_without_sheet_names(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        path = exp.export(ROWS, str(tmp_path / 'out.xlsx'))
        assert os.path.exists(path)

    def test_creates_xlsx_with_sheet_names(self, tmp_path):
        import openpyxl
        exp = ExcelExporter(str(tmp_path))
        path = exp.export(ROWS, str(tmp_path / 'out.xlsx'), sheet_names=['CASH', 'PROCE'])
        wb = openpyxl.load_workbook(path)
        assert 'CASH' in wb.sheetnames
        assert 'PROCE' in wb.sheetnames
        assert 'combined' in wb.sheetnames

    def test_sheet_only_contains_matching_rows(self, tmp_path):
        import openpyxl
        exp = ExcelExporter(str(tmp_path))
        path = exp.export(ROWS, str(tmp_path / 'out.xlsx'), sheet_names=['CASH', 'PROCE'])
        wb = openpyxl.load_workbook(path)
        cash_rows = list(wb['CASH'].iter_rows(min_row=2, values_only=True))
        # 2 contratos CASH en ROWS
        assert len(cash_rows) == 2


class TestExcelExporterHelpers:
    def test_sanitize_sheet_name_truncates_to_31(self):
        long_name = 'A' * 40
        result = ExcelExporter._sanitize_sheet_name(long_name)
        assert len(result) == 31

    def test_sanitize_sheet_name_removes_invalid_chars(self):
        assert ExcelExporter._sanitize_sheet_name('hello/world') == 'helloworld'
        assert ExcelExporter._sanitize_sheet_name('a[b]c') == 'abc'

    def test_sanitize_sheet_name_fallback_for_empty(self):
        assert ExcelExporter._sanitize_sheet_name('') == 'sheet'
        assert ExcelExporter._sanitize_sheet_name('////') == 'sheet'

    def test_truncate_value_short_string_unchanged(self):
        assert ExcelExporter._truncate_value('hello') == 'hello'

    def test_truncate_value_long_string_truncated(self):
        long = 'x' * 40000
        result = ExcelExporter._truncate_value(long)
        assert len(result) == ExcelExporter.MAX_CELL_LENGTH
        assert result.endswith('...')

    def test_truncate_value_non_string_unchanged(self):
        assert ExcelExporter._truncate_value(42) == 42
        assert ExcelExporter._truncate_value(None) is None


# ── ExportOrchestrator ────────────────────────────────────────────────────────

class TestExportOrchestratorPartial:
    def test_no_op_when_partial_export_false(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path), partial_export=False, save_every=10)
        asyncio.run(orch.export_partial(ROWS, 'test'))
        # no debe haber creado archivos
        files = list(tmp_path.glob('partial_*'))
        assert files == []

    def test_no_op_when_save_every_zero(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path), partial_export=True, save_every=0)
        asyncio.run(orch.export_partial(ROWS, 'test'))
        assert list(tmp_path.glob('partial_*')) == []

    def test_no_op_when_rows_empty(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path), partial_export=True, save_every=10)
        asyncio.run(orch.export_partial([], 'test'))
        assert list(tmp_path.glob('partial_*')) == []

    def test_writes_csv_when_partial_export_active(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(
            exp, str(tmp_path), partial_export=True, partial_format='csv', save_every=1
        )
        asyncio.run(orch.export_partial(ROWS, 'worldclass_CASH'))
        files = list(tmp_path.glob('partial_*.csv'))
        assert len(files) == 1


class TestExportOrchestratorSiteReports:
    def test_creates_directory_structure(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path))
        asyncio.run(orch.export_site_reports(
            sitio=SITIO, rows=ROWS, mode='worldclass', export_csv=True, export_xlsx=False
        ))
        # debe crear <output>/worldclass/<sede-slug>/reports/
        wcg_dir = tmp_path / 'worldclass' / 'wcg-guayaquil' / 'reports'
        wc_dir  = tmp_path / 'worldclass' / 'wc-santo-domingo' / 'reports'
        assert wcg_dir.exists()
        assert wc_dir.exists()

    def test_creates_one_csv_per_sede_estado_group(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path))
        asyncio.run(orch.export_site_reports(
            sitio=SITIO, rows=ROWS, mode='worldclass', export_csv=True, export_xlsx=False
        ))
        # grupos: (WCG, CASH), (WCG, PROCE), (WC, CASH) → 3 csvs
        all_csvs = list((tmp_path / 'worldclass').rglob('*.csv'))
        assert len(all_csvs) == 3

    def test_no_files_when_rows_empty(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path))
        asyncio.run(orch.export_site_reports(
            sitio=SITIO, rows=[], mode='worldclass', export_csv=True, export_xlsx=False
        ))
        assert list(tmp_path.rglob('*.csv')) == []

    def test_sede_override_used_when_row_has_no_sede(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path))
        rows_no_sede = [{'Estado_Contrato': 'CASH', 'Nombre_Titular': 'Test'}]
        asyncio.run(orch.export_site_reports(
            sitio=SITIO, rows=rows_no_sede, mode='worldclass',
            sede_override='WCG - GUAYAQUIL', export_csv=True
        ))
        wcg_dir = tmp_path / 'worldclass' / 'wcg-guayaquil' / 'reports'
        assert wcg_dir.exists()

    def test_xlsx_created_when_requested(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path))
        asyncio.run(orch.export_site_reports(
            sitio=SITIO, rows=ROWS[:1], mode='worldclass',
            export_csv=False, export_xlsx=True
        ))
        xlsxs = list((tmp_path / 'worldclass').rglob('*.xlsx'))
        assert len(xlsxs) == 1


class TestExportOrchestratorCombined:
    def test_export_combined_csv(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path))
        saved = asyncio.run(orch.export_combined(
            rows=ROWS, mode='todos', filename='contratos_todos.xlsx',
            export_csv=True, export_xlsx=False
        ))
        assert len(saved) == 1
        assert saved[0].endswith('.csv')
        assert os.path.exists(saved[0])

    def test_export_combined_xlsx(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path))
        saved = asyncio.run(orch.export_combined(
            rows=ROWS, mode='todos', filename='contratos_todos.xlsx',
            export_csv=False, export_xlsx=True
        ))
        assert len(saved) == 1
        assert saved[0].endswith('.xlsx')
        assert os.path.exists(saved[0])

    def test_export_combined_returns_empty_for_no_rows(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path))
        saved = asyncio.run(orch.export_combined(
            rows=[], mode='todos', filename='contratos_todos.xlsx',
            export_csv=True, export_xlsx=True
        ))
        assert saved == []

    def test_export_combined_both_formats(self, tmp_path):
        exp = ExcelExporter(str(tmp_path))
        orch = ExportOrchestrator(exp, str(tmp_path))
        saved = asyncio.run(orch.export_combined(
            rows=ROWS, mode='todos', filename='contratos_todos.xlsx',
            export_csv=True, export_xlsx=True
        ))
        assert len(saved) == 2
